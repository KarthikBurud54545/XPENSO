from django.shortcuts import render,HttpResponse,redirect, get_object_or_404
print("DEBUG: views.py - Starting imports")
from django.contrib import messages
from django.contrib.auth import authenticate ,logout
from django.contrib.auth import login as dj_login
from django.contrib.auth.models import User
from .models import (
    Addmoney_info, 
    UserProfile, 
    Budget, 
    Investment, 
    INVESTMENT_TYPE_CHOICES, 
    INVESTMENT_STATUS_CHOICES,
    SELECT_CATEGORY_CHOICES
)
from django.contrib.sessions.models import Session
from django.core.paginator import Paginator, EmptyPage , PageNotAnInteger
from django.db.models import Sum, Avg, Count
from django.http import JsonResponse
import datetime
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from decimal import Decimal
import json
from django.views.decorators.http import require_http_methods, require_GET
import csv
from django.http import HttpResponse
print("DEBUG: views.py - Importing reportlab")
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus.tables import Table
print("DEBUG: views.py - Importing openpyxl")
from openpyxl import Workbook
from django.views.decorators.csrf import csrf_exempt
print("DEBUG: views.py - Importing numpy")
import numpy as np
print("DEBUG: views.py - Importing sklearn")
from sklearn.linear_model import LinearRegression
print("DEBUG: views.py - Importing timedelta")
from datetime import timedelta
print("DEBUG: views.py - Importing calendar")
import calendar
print("DEBUG: views.py - Importing utils")
from .utils import get_real_time_stock_price, get_market_news, calculate_portfolio_metrics, get_portfolio_recommendations
print("DEBUG: views.py - Importing StreamingHttpResponse")
from django.http import StreamingHttpResponse
print("DEBUG: views.py - Importing io")
import io
print("DEBUG: views.py - Importing defaultdict")
from collections import defaultdict
print("DEBUG: views.py - All imports finished")

# Create your views here.
def home(request):
    if request.session.has_key('is_logged'):
        return redirect('/index')
    return render(request,'home/login.html')
   # return HttpResponse('This is home')
def index(request):
    if request.session.has_key('is_logged'):
        user_id = request.session["user_id"]
        user = User.objects.get(id=user_id)
        addmoney_info = Addmoney_info.objects.filter(user=user).order_by('-Date')
        paginator = Paginator(addmoney_info, 4)
        page_number = request.GET.get('page')
        page_obj = Paginator.get_page(paginator, page_number)

        # Calculate totals
        total_income = addmoney_info.filter(add_money='Income').aggregate(Sum('quantity'))['quantity__sum'] or 0
        total_expenses = addmoney_info.filter(add_money='Expense').aggregate(Sum('quantity'))['quantity__sum'] or 0
        total_balance = total_income - total_expenses

        # Get budget information
        budgets = Budget.objects.filter(user=user)
        for budget in budgets:
            # Calculate spent amount for each budget category
            spent = addmoney_info.filter(
                Category=budget.category,
                add_money='Expense',
                Date__month=datetime.datetime.now().month
            ).aggregate(Sum('quantity'))['quantity__sum'] or 0
            budget.spent = spent

        # Get investment information - using purchase_date for ordering
        investments = Investment.objects.filter(user=user).order_by('-purchase_date')[:5]

        # Get monthly spending breakdown
        current_month = timezone.now().month
        monthly_expenses = addmoney_info.filter(
            add_money='Expense',
            Date__month=current_month
        ).values('Category').annotate(
            total=Sum('quantity')
        ).order_by('-total')

        # Get yearly expense breakdown for pie chart
        current_year = timezone.now().year
        yearly_expenses = addmoney_info.filter(
            add_money='Expense',
            Date__year=current_year
        ).values('Category').annotate(
            total=Sum('quantity')
        ).order_by('-total')

        # Prepare data for the pie chart
        expense_categories = [item['Category'] for item in yearly_expenses]
        expense_amounts = [float(item['total']) for item in yearly_expenses]
        
        # Generate colors for the pie chart
        colors = [
            '#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF',
            '#FF9F40', '#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0'
        ][:len(expense_categories)]

        expense_data_combined = zip(expense_categories, expense_amounts, colors)

        # JSON serialize the chart data
        chart_data = {
            'categories': expense_categories,
            'amounts': expense_amounts,
            'colors': colors
        }

        context = {
            'page_obj': page_obj,
            'total_income': total_income,
            'total_expenses': total_expenses,
            'total_balance': total_balance,
            'budgets': budgets,
            'investments': investments,
            'monthly_expenses': monthly_expenses,
            'chart_data': json.dumps(chart_data),
            'current_year': current_year,
            'expense_amounts': expense_amounts,
            'expense_categories': expense_categories,
            'colors': colors,
            'expense_data_combined': expense_data_combined,
        }
        return render(request, 'home/index.html', context)
    return redirect('home')
    #return HttpResponse('This is blog')
def register(request):
    return render(request,'home/register.html')
    #return HttpResponse('This is blog')
def password(request):
    return render(request,'home/password.html')

def charts(request):
    return render(request,'home/charts.html')
def search(request):
    if request.session.has_key('is_logged'):
        user_id = request.session["user_id"]
        user = User.objects.get(id=user_id)
        
        # Get filter parameters
        fromdate = request.GET.get('fromdate')
        todate = request.GET.get('todate')
        category = request.GET.get('category')
        transaction_type = request.GET.get('type')
        
        # Start with all user's transactions
        addmoney = Addmoney_info.objects.filter(user=user)
        
        # Apply filters
        if fromdate and todate:
            addmoney = addmoney.filter(Date__range=[fromdate, todate])
        
        if category:
            addmoney = addmoney.filter(Category=category)
            
        if transaction_type:
            addmoney = addmoney.filter(add_money=transaction_type)
            
        # Order by date
        addmoney = addmoney.order_by('-Date')
        
        # Calculate filtered totals
        total_income = addmoney.filter(add_money='Income').aggregate(Sum('quantity'))['quantity__sum'] or 0
        total_expenses = addmoney.filter(add_money='Expense').aggregate(Sum('quantity'))['quantity__sum'] or 0
        net_balance = total_income - total_expenses
        
        # Get categories for filter dropdown
        categories = Addmoney_info.objects.filter(user=user).values_list('Category', flat=True).distinct()
        
        context = {
            'addmoney': addmoney,
            'total_income': total_income,
            'total_expenses': total_expenses,
            'net_balance': net_balance,
            'categories': categories,
            'fromdate': fromdate,
            'todate': todate,
            'selected_category': category,
            'selected_type': transaction_type,
        }
        
        return render(request, 'home/tables.html', context)
    return redirect('home')

def tables(request):
    if request.session.has_key('is_logged'):
        user_id = request.session["user_id"]
        user = User.objects.get(id=user_id)
        
        # Get all transactions for the user
        addmoney = Addmoney_info.objects.filter(user=user).order_by('-Date')
        
        # Calculate totals
        total_income = addmoney.filter(add_money='Income').aggregate(Sum('quantity'))['quantity__sum'] or 0
        total_expenses = addmoney.filter(add_money='Expense').aggregate(Sum('quantity'))['quantity__sum'] or 0
        net_balance = total_income - total_expenses
        
        # Get unique categories for filter dropdown
        categories = addmoney.values_list('Category', flat=True).distinct()
        
        context = {
            'addmoney': addmoney,
            'total_income': total_income,
            'total_expenses': total_expenses,
            'net_balance': net_balance,
            'categories': categories,
        }
        
        return render(request, 'home/tables.html', context)
    return redirect('home')

def addmoney(request):
    return render(request,'home/addmoney.html')

def profile(request):
    if request.session.has_key('is_logged'):
        return render(request,'home/profile.html')
    return redirect('/home')

def profile_edit(request,id):
    if request.session.has_key('is_logged'):
        add = User.objects.get(id=id)
        # user_id = request.session["user_id"]
        # user1 = User.objects.get(id=user_id)
        return render(request,'home/profile_edit.html',{'add':add})
    return redirect("/home")

def profile_update(request,id):
    if request.session.has_key('is_logged'):
        if request.method == "POST":
            user = User.objects.get(id=id)
            user.first_name = request.POST["fname"]
            user.last_name = request.POST["lname"]
            user.email = request.POST["email"]
            user.userprofile.Savings = request.POST["Savings"]
            user.userprofile.income = request.POST["income"]
            user.userprofile.profession = request.POST["profession"]
            # Handle profile image upload
            if 'file' in request.FILES:
                user.userprofile.image = request.FILES['file']
            user.userprofile.save()
            user.save()
            return redirect("/profile")
    return redirect("/home")   

def handleSignup(request):
    if request.method =='POST':
            # get the post parameters
            uname = request.POST["uname"]
            fname=request.POST["fname"]
            lname=request.POST["lname"]
            email = request.POST["email"]
            profession = request.POST['profession']
            Savings = request.POST['Savings']
            income = request.POST['income']
            pass1 = request.POST["pass1"]
            pass2 = request.POST["pass2"]
            profile = UserProfile(Savings = Savings,profession=profession,income=income)
            # check for errors in input
            if request.method == 'POST':
                try:
                    user_exists = User.objects.get(username=request.POST['uname'])
                    messages.error(request," Username already taken, Try something else!!!")
                    return redirect("/register")    
                except User.DoesNotExist:
                    if len(uname)>15:
                        messages.error(request," Username must be max 15 characters, Please try again")
                        return redirect("/register")
            
                    if not uname.isalnum():
                        messages.error(request," Username should only contain letters and numbers, Please try again")
                        return redirect("/register")
            
                    if pass1 != pass2:
                        messages.error(request," Password do not match, Please try again")
                        return redirect("/register")
            
            # create the user
            user = User.objects.create_user(uname, email, pass1)
            user.first_name=fname
            user.last_name=lname
            user.email = email
            # profile = UserProfile.objects.all()

            user.save()
            # p1=profile.save(commit=False)
            profile.user = user
            profile.save()
            messages.success(request," Your account has been successfully created")
            return redirect("/")
    else:
        return HttpResponse('404 - NOT FOUND ')
    return redirect('/login')

def handlelogin(request):
    if request.method =='POST':
        # get the post parameters
        loginuname = request.POST["loginuname"]
        loginpassword1=request.POST["loginpassword1"]
        user = authenticate(username=loginuname, password=loginpassword1)
        if user is not None:
            dj_login(request, user)
            request.session['is_logged'] = True
            user = request.user.id 
            request.session["user_id"] = user
            messages.success(request, " Successfully logged in")
            return redirect('/index')
        else:
            messages.error(request," Invalid Credentials, Please try again")  
            return redirect("/")  
    return HttpResponse('404-not found')
def handleLogout(request):
        del request.session['is_logged']
        del request.session["user_id"] 
        logout(request)
        messages.success(request, " Successfully logged out")
        return redirect('home')

#add money form
def addmoney_submission(request):
    if request.session.has_key('is_logged'):
        if request.method == "POST":
            user_id = request.session["user_id"]
            user1 = User.objects.get(id=user_id)
            addmoney_info1 = Addmoney_info.objects.filter(user=user1).order_by('-Date')
            add_money = request.POST["add_money"]
            quantity = request.POST["quantity"]
            Date = request.POST["Date"]
            Category = request.POST["Category"]
            add = Addmoney_info(user = user1,add_money=add_money,quantity=quantity,Date = Date,Category= Category)
            add.save()
            paginator = Paginator(addmoney_info1, 4)
            page_number = request.GET.get('page')
            page_obj = Paginator.get_page(paginator,page_number)
            
            # Calculate totals (same as index function)
            total_income = addmoney_info1.filter(add_money='Income').aggregate(Sum('quantity'))['quantity__sum'] or 0
            total_expenses = addmoney_info1.filter(add_money='Expense').aggregate(Sum('quantity'))['quantity__sum'] or 0
            total_balance = total_income - total_expenses

            # Get budget information
            budgets = Budget.objects.filter(user=user1)
            for budget in budgets:
                # Calculate spent amount for each budget category
                spent = addmoney_info1.filter(
                    Category=budget.category,
                    add_money='Expense',
                    Date__month=datetime.datetime.now().month
                ).aggregate(Sum('quantity'))['quantity__sum'] or 0
                budget.spent = spent

            # Get investment information - using purchase_date for ordering
            investments = Investment.objects.filter(user=user1).order_by('-purchase_date')[:5]

            # Get monthly spending breakdown
            current_month = timezone.now().month
            monthly_expenses = addmoney_info1.filter(
                add_money='Expense',
                Date__month=current_month
            ).values('Category').annotate(
                total=Sum('quantity')
            ).order_by('-total')

            # Get yearly expense breakdown for pie chart
            current_year = timezone.now().year
            yearly_expenses = addmoney_info1.filter(
                add_money='Expense',
                Date__year=current_year
            ).values('Category').annotate(
                total=Sum('quantity')
            ).order_by('-total')

            # Prepare data for the pie chart
            expense_categories = [item['Category'] for item in yearly_expenses]
            expense_amounts = [float(item['total']) for item in yearly_expenses]
            
            # Generate colors for the pie chart
            colors = [
                '#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0', '#9966FF',
                '#FF9F40', '#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0'
            ][:len(expense_categories)]

            expense_data_combined = zip(expense_categories, expense_amounts, colors)

            # JSON serialize the chart data
            chart_data = {
                'categories': expense_categories,
                'amounts': expense_amounts,
                'colors': colors
            }

            context = {
                'page_obj': page_obj,
                'total_income': total_income,
                'total_expenses': total_expenses,
                'total_balance': total_balance,
                'budgets': budgets,
                'investments': investments,
                'monthly_expenses': monthly_expenses,
                'chart_data': json.dumps(chart_data),
                'current_year': current_year,
                'expense_amounts': expense_amounts,
                'expense_categories': expense_categories,
                'colors': colors,
                'expense_data_combined': expense_data_combined,
            }
            return render(request,'home/index.html',context)
    return redirect('/index')

def addmoney_update(request,id):
    if request.session.has_key('is_logged'):
        if request.method == "POST":
            add  = Addmoney_info.objects.get(id=id)
            add .add_money = request.POST["add_money"]
            add.quantity = request.POST["quantity"]
            add.Date = request.POST["Date"]
            add.Category = request.POST["Category"]
            add .save()
            return redirect("/index")
    return redirect("/home")        

def expense_edit(request,id):
    if request.session.has_key('is_logged'):
        addmoney_info = Addmoney_info.objects.get(id=id)
        user_id = request.session["user_id"]
        user1 = User.objects.get(id=user_id)
        return render(request,'home/expense_edit.html',{'addmoney_info':addmoney_info})
    return redirect("/home")  

def expense_delete(request,id):
    if request.session.has_key('is_logged'):
        addmoney_info = Addmoney_info.objects.get(id=id)
        addmoney_info.delete()
        return redirect("/index")
    return redirect("/home")  

def expense_month(request):
    todays_date = datetime.date.today()
    one_month_ago = todays_date-datetime.timedelta(days=30)
    user_id = request.session["user_id"]
    user1 = User.objects.get(id=user_id)
    addmoney = Addmoney_info.objects.filter(user = user1,Date__gte=one_month_ago,Date__lte=todays_date)
    finalrep ={}

    def get_Category(addmoney_info):
        # if addmoney_info.add_money=="Expense":
        return addmoney_info.Category    
    Category_list = list(set(map(get_Category,addmoney)))

    def get_expense_category_amount(Category,add_money):
        quantity = 0 
        filtered_by_category = addmoney.filter(Category = Category,add_money="Expense") 
        for item in filtered_by_category:
            quantity+=item.quantity
        return quantity

    for x in addmoney:
        for y in Category_list:
            finalrep[y]= get_expense_category_amount(y,"Expense")

    return JsonResponse({'expense_category_data': finalrep}, safe=False)


def stats(request):
    if request.session.has_key('is_logged'):
        user_id = request.session["user_id"]
        user1 = User.objects.get(id=user_id)
        
        # Get current month's data
        today = datetime.date.today()
        first_day = today.replace(day=1)
        last_day = (first_day + datetime.timedelta(days=32)).replace(day=1) - datetime.timedelta(days=1)
        
        # Get monthly transactions
        monthly_transactions = Addmoney_info.objects.filter(
            user=user1,
            Date__gte=first_day,
            Date__lte=last_day
        )
        
        # Calculate totals
        total_income = monthly_transactions.filter(add_money='Income').aggregate(Sum('quantity'))['quantity__sum'] or 0
        total_expenses = monthly_transactions.filter(add_money='Expense').aggregate(Sum('quantity'))['quantity__sum'] or 0
        total_balance = (user1.userprofile.Savings or 0) + total_income - total_expenses
        
        # Get category-wise expenses
        category_expenses = monthly_transactions.filter(
            add_money='Expense'
        ).values('Category').annotate(
            total=Sum('quantity')
        ).order_by('-total')
        
        # Get daily expense trend
        daily_expenses = monthly_transactions.filter(
            add_money='Expense'
        ).values('Date').annotate(
            total=Sum('quantity')
        ).order_by('Date')
        
        # Get budget comparison with better error handling
        budgets = Budget.objects.filter(user=user1)
        budget_data = []
        
        if budgets.exists():
            for budget in budgets:
                spent = monthly_transactions.filter(
                    Category=budget.category,
                    add_money='Expense'
                ).aggregate(Sum('quantity'))['quantity__sum'] or 0
                
                # Convert to float for consistent decimal handling
                spent = float(spent)
                budget_amount = float(budget.amount)
                
                budget_data.append({
                    'category': budget.category,
                    'budget': budget_amount,
                    'spent': spent,
                    'remaining': max(0, budget_amount - spent)  # Ensure remaining doesn't go negative
                })
            
            # Sort budgets by percentage spent
            budget_data.sort(key=lambda x: (x['spent'] / x['budget']) if x['budget'] > 0 else 0, reverse=True)
        
        context = {
            'total_income': total_income,
            'total_expenses': total_expenses,
            'total_balance': total_balance,
            'category_expenses': category_expenses,
            'daily_expenses': daily_expenses,
            'budget_data': budget_data,
            'month_name': today.strftime('%B'),
            'year': today.year,
            'has_budgets': bool(budget_data)  # Add flag to check if budgets exist
        }
        
        return render(request, 'home/stats.html', context)
    return redirect('home')

def expense_month_data(request):
    if request.session.has_key('is_logged'):
        user_id = request.session["user_id"]
        user = User.objects.get(id=user_id)
        
        # Get current month's data
        today = datetime.date.today()
        first_day = today.replace(day=1)
        last_day = (first_day + datetime.timedelta(days=32)).replace(day=1) - datetime.timedelta(days=1)
        
        # Get daily expense data
        daily_expenses = Addmoney_info.objects.filter(
            user=user,
            Date__gte=first_day,
            Date__lte=last_day,
            add_money='Expense'
        ).values('Date').annotate(
            total=Sum('quantity')
        ).order_by('Date')
        
        # Get category data
        category_expenses = Addmoney_info.objects.filter(
            user=user,
            Date__gte=first_day,
            Date__lte=last_day,
            add_money='Expense'
        ).values('Category').annotate(
            total=Sum('quantity')
        ).order_by('-total')
        
        # Format data for charts
        dates = [expense['Date'].strftime('%d %b') for expense in daily_expenses]
        amounts = [float(expense['total']) for expense in daily_expenses]
        
        categories = [expense['Category'] for expense in category_expenses]
        category_amounts = [float(expense['total']) for expense in category_expenses]
        
        data = {
            'dates': dates,
            'daily_amounts': amounts,
            'categories': categories,
            'category_amounts': category_amounts
        }
        
        return JsonResponse(data)
    return JsonResponse({'error': 'Not logged in'}, status=401)

def expense_week(request):
    todays_date = datetime.date.today()
    one_week_ago = todays_date-datetime.timedelta(days=7)
    user_id = request.session["user_id"]
    user1 = User.objects.get(id=user_id)
    addmoney = Addmoney_info.objects.filter(user = user1,Date__gte=one_week_ago,Date__lte=todays_date)
    finalrep ={}

    def get_Category(addmoney_info):
        return addmoney_info.Category
    Category_list = list(set(map(get_Category,addmoney)))


    def get_expense_category_amount(Category,add_money):
        quantity = 0 
        filtered_by_category = addmoney.filter(Category = Category,add_money="Expense") 
        for item in filtered_by_category:
            quantity+=item.quantity
        return quantity

    for x in addmoney:
        for y in Category_list:
            finalrep[y]= get_expense_category_amount(y,"Expense")

    return JsonResponse({'expense_category_data': finalrep}, safe=False)
    
@login_required
def weekly(request):
    # Get current user
    user = request.user
    
    # Get current week's start and end dates
    today = datetime.date.today()
    start_of_week = today - datetime.timedelta(days=today.weekday())
    end_of_week = start_of_week + datetime.timedelta(days=6)

    # Get transactions for current week
    current_week_transactions = Addmoney_info.objects.filter(
        user=user,
        Date__range=[start_of_week, end_of_week]
    )

    # Calculate totals for current week
    total_income = current_week_transactions.filter(
        add_money='Income'
    ).aggregate(Sum('quantity'))['quantity__sum'] or 0

    total_expenses = current_week_transactions.filter(
        add_money='Expense'
    ).aggregate(Sum('quantity'))['quantity__sum'] or 0

    total_balance = float(total_income) - float(total_expenses)

    # Get last week's data for comparison
    last_week_start = start_of_week - datetime.timedelta(days=7)
    last_week_end = end_of_week - datetime.timedelta(days=7)

    last_week_transactions = Addmoney_info.objects.filter(
        user=user,
        Date__range=[last_week_start, last_week_end]
    )

    last_week_income = last_week_transactions.filter(
        add_money='Income'
    ).aggregate(Sum('quantity'))['quantity__sum'] or 0

    last_week_expenses = last_week_transactions.filter(
        add_money='Expense'
    ).aggregate(Sum('quantity'))['quantity__sum'] or 0

    last_week_balance = float(last_week_income) - float(last_week_expenses)

    # Get category breakdown
    categories = []
    expense_categories = current_week_transactions.filter(
        add_money='Expense'
    ).values('Category').annotate(
        total=Sum('quantity')
    ).order_by('-total')

    total_expense_amount = sum(cat['total'] for cat in expense_categories)

    for category in expense_categories:
        cat_name = category['Category']
        amount = float(category['total'])
        percentage = round((amount / total_expense_amount * 100) if total_expense_amount > 0 else 0, 1)
        
        # Calculate trend (% change from last week)
        last_week_amount = last_week_transactions.filter(
            add_money='Expense',
            Category=cat_name
        ).aggregate(Sum('quantity'))['quantity__sum'] or 0
        
        trend = 0
        if float(last_week_amount) > 0:
            trend = round(((amount - float(last_week_amount)) / float(last_week_amount) * 100), 1)

        categories.append({
            'name': cat_name,
            'amount': amount,
            'percentage': percentage,
            'trend': trend
        })

    # Get daily spending data
    daily_spending = {}
    for i in range(7):
        date = start_of_week + datetime.timedelta(days=i)
        amount = current_week_transactions.filter(
            Date=date,
            add_money='Expense'
        ).aggregate(Sum('quantity'))['quantity__sum'] or 0
        daily_spending[date.strftime('%a')] = float(amount)

    # Prepare chart data
    chart_data = {
        'categories': categories,
        'dailySpending': daily_spending
    }

    context = {
        'addmoney_info': current_week_transactions,
        'total_income': total_income,
        'total_expenses': total_expenses,
        'total_balance': total_balance,
        'categories': categories,
        'previous_balance': last_week_balance,
        'last_week_income': last_week_income,
        'last_week_expenses': last_week_expenses,
        'start_date': start_of_week,
        'end_date': end_of_week,
        'chart_data': json.dumps(chart_data, default=str)
    }

    return render(request, 'home/weekly.html', context)

def check(request):
    if request.method == 'POST':
        user_exists = User.objects.filter(email=request.POST['email'])
        messages.error(request,"Email not registered, TRY AGAIN!!!")
        return redirect("/reset_password")

def info_year(request):
    if request.session.has_key('is_logged'):
        user_id = request.session["user_id"]
        user = User.objects.get(id=user_id)
        
        # Get current year's data
        current_year = datetime.datetime.now().year
        yearly_expenses = Addmoney_info.objects.filter(
            user=user,
            Date__year=current_year,
            add_money='Expense'
        ).values('Category').annotate(
            total=Sum('quantity')
        ).order_by('-total')

        # Create the expense category data dictionary
        expense_category_data = {
            expense['Category']: float(expense['total'])
            for expense in yearly_expenses
        }

        return JsonResponse({'expense_category_data': expense_category_data}, safe=False)
    return JsonResponse({'error': 'Not logged in'}, status=401)

def info(request):
    if request.session.has_key('is_logged'):
        user_id = request.session["user_id"]
        user = User.objects.get(id=user_id)
        
        # Get current year's data
        current_year = datetime.datetime.now().year
        yearly_transactions = Addmoney_info.objects.filter(
            user=user,
            Date__year=current_year
        )
        
        # Get previous year's data for comparison
        previous_year = current_year - 1
        previous_transactions = Addmoney_info.objects.filter(
            user=user,
            Date__year=previous_year
        )
        
        # Calculate current year totals
        total_income = yearly_transactions.filter(add_money='Income').aggregate(Sum('quantity'))['quantity__sum'] or 0
        total_expenses = yearly_transactions.filter(add_money='Expense').aggregate(Sum('quantity'))['quantity__sum'] or 0
        total_balance = (user.userprofile.Savings or 0) + total_income - total_expenses
        
        # Calculate previous year totals for comparison
        previous_income = previous_transactions.filter(add_money='Income').aggregate(Sum('quantity'))['quantity__sum'] or 0
        previous_expenses = previous_transactions.filter(add_money='Expense').aggregate(Sum('quantity'))['quantity__sum'] or 0
        previous_balance = previous_income - previous_expenses
        
        # Calculate net savings and savings rate
        net_savings = total_income - total_expenses
        savings_rate = (net_savings / total_income * 100) if total_income > 0 else 0
        
        # Get top expenses with trend
        top_expenses = []
        category_expenses = yearly_transactions.filter(
            add_money='Expense'
        ).values('Category').annotate(
            amount=Sum('quantity')
        ).order_by('-amount')[:5]
        
        total_expense_amount = sum(expense['amount'] for expense in category_expenses)
        
        for expense in category_expenses:
            # Calculate previous year amount for this category
            previous_amount = previous_transactions.filter(
                add_money='Expense',
                Category=expense['Category']
            ).aggregate(Sum('quantity'))['quantity__sum'] or 0
            
            # Calculate trend percentage
            if previous_amount > 0:
                trend = ((expense['amount'] - previous_amount) / previous_amount * 100)
            else:
                trend = 100 if expense['amount'] > 0 else 0
            
            top_expenses.append({
                'category': expense['Category'],
                'amount': expense['amount'],
                'percentage': (expense['amount'] / total_expense_amount * 100) if total_expense_amount > 0 else 0,
                'trend': trend
            })
        
        context = {
            'year': current_year,
            'total_income': total_income,
            'total_expenses': total_expenses,
            'total_balance': total_balance,
            'previous_income': previous_income,
            'previous_expenses': previous_expenses,
            'previous_balance': previous_balance,
            'net_savings': net_savings,
            'savings_rate': savings_rate,
            'top_expenses': top_expenses
        }
        
        return render(request, 'home/info.html', context)
    return redirect('home')

def login_page(request):
    return render(request, 'home/login.html')

@login_required
def budgets_view(request):
    from .models import Budget, Addmoney_info
    import datetime
    from django.db.models import Sum
    # Handle new budget submission
    if request.method == "POST":
        name = request.POST.get("budgetName")
        amount = request.POST.get("budgetAmount")
        category = request.POST.get("budgetCategory")
        period = request.POST.get("budgetPeriod")
        if name and amount and category and period:
            Budget.objects.create(
                user=request.user,
                name=name,
                amount=amount,
                category=category,
                period=period
            )
    user_budgets = Budget.objects.filter(user=request.user)
    # Calculate spent for each budget
    for budget in user_budgets:
        spent = Addmoney_info.objects.filter(
            user=request.user,
            Category=budget.category,
            add_money='Expense',
            Date__month=datetime.datetime.now().month
        ).aggregate(Sum('quantity'))['quantity__sum'] or 0
        budget.spent = spent
    # Calculate summary
    total_budget = sum(float(b.amount) for b in user_budgets)
    total_spent = sum(float(getattr(b, 'spent', 0)) for b in user_budgets)
    total_remaining = total_budget - total_spent
    # Aggregate by category for chart
    category_totals = defaultdict(lambda: {'budget': 0, 'spent': 0})
    for b in user_budgets:
        category_totals[b.name]['budget'] += float(b.amount)
        category_totals[b.name]['spent'] += abs(float(getattr(b, 'spent', 0)))
    chart_labels = list(category_totals.keys())
    chart_budget = [v['budget'] for v in category_totals.values()]
    chart_spent = [v['spent'] for v in category_totals.values()]
    context = {
        'budgets': user_budgets,
        'total_budget': total_budget,
        'total_spent': total_spent,
        'total_remaining': total_remaining,
        'chart_labels': json.dumps(chart_labels),
        'chart_spent': json.dumps(chart_spent),
        'chart_budget': json.dumps(chart_budget),
    }
    return render(request, 'home/budgets.html', context)

def investments(request):
    if request.session.has_key('is_logged'):
        user_id = request.session["user_id"]
        user = User.objects.get(id=user_id)
        investments = Investment.objects.filter(user=user)
        
        # Calculate total investment metrics
        total_invested = sum(inv.amount_invested for inv in investments)
        total_current = sum(inv.current_value for inv in investments)
        total_returns = total_current - total_invested
        total_returns_percentage = (total_returns / total_invested * 100) if total_invested > 0 else 0
        
        context = {
            'investments': investments,
            'total_invested': total_invested,
            'current_value': total_current,
            'total_returns': total_returns,
            'return_rate': total_returns_percentage,
        }
        return render(request, 'home/investments.html', context)
    return redirect('home')

def add_investment(request):
    if request.session.has_key('is_logged'):
        if request.method == "POST":
            user_id = request.session["user_id"]
            user = User.objects.get(id=user_id)
            
            try:
                investment = Investment(
                    user=user,
                    investment_type=request.POST["investment_type"],
                    name=request.POST["name"],
                    amount_invested=Decimal(request.POST["amount_invested"]),
                    current_value=Decimal(request.POST["current_value"]),
                    purchase_date=request.POST["purchase_date"],
                    maturity_date=request.POST.get("maturity_date") or None,
                    status=request.POST.get("status", "Active"),
                    notes=request.POST.get("notes", "")
                )
                investment.save()
                messages.success(request, "Investment added successfully!")
                return redirect('investments')
            except Exception as e:
                messages.error(request, f"Error adding investment: {str(e)}")
                return redirect('add_investment')
        
        # Pass the investment choices to the template
        context = {
            'investment_types': INVESTMENT_TYPE_CHOICES,
            'investment_statuses': INVESTMENT_STATUS_CHOICES
        }
        return render(request, 'home/add_investment.html', context)
    return redirect('home')

def update_investment(request, id):
    if request.session.has_key('is_logged'):
        if request.method == "POST":
            try:
                investment = Investment.objects.get(id=id, user_id=request.session["user_id"])
                investment.investment_type = request.POST["investment_type"]
                investment.name = request.POST["name"]
                investment.amount_invested = Decimal(request.POST["amount_invested"])
                investment.current_value = Decimal(request.POST["current_value"])
                investment.purchase_date = request.POST["purchase_date"]
                investment.maturity_date = request.POST.get("maturity_date") or None
                investment.status = request.POST.get("status", "Active")
                investment.notes = request.POST.get("notes", "")
                investment.save()
                messages.success(request, "Investment updated successfully!")
                return redirect('investments')
            except Exception as e:
                messages.error(request, f"Error updating investment: {str(e)}")
                return redirect('update_investment', id=id)
                
        investment = Investment.objects.get(id=id, user_id=request.session["user_id"])
        context = {
            'investment': investment,
            'investment_types': INVESTMENT_TYPE_CHOICES,
            'investment_statuses': INVESTMENT_STATUS_CHOICES
        }
        return render(request, 'home/update_investment.html', context)
    return redirect('home')

def view_investment(request, id):
    if request.session.has_key('is_logged'):
        investment = get_object_or_404(Investment, id=id, user_id=request.session["user_id"])
        context = {
            'investment': investment
        }
        return render(request, 'home/investment_detail.html', context)
    return redirect('home')

def delete_investment(request, id):
    investment = get_object_or_404(Investment, id=id, user=request.user)
    investment.delete()
    messages.success(request, 'Investment deleted successfully!', extra_tags='success_modal')
    return redirect('investments')

def get_market_data(request):
    """API endpoint for real-time market data"""
    try:
        # Get S&P 500 data
        sp500 = yf.Ticker('^GSPC')
        sp500_price = sp500.info.get('regularMarketPrice', 'N/A')
        sp500_change = sp500.info.get('regularMarketChangePercent', 0)
        
        # Get USD/INR rate
        usdinr = yf.Ticker('INR=X')
        usdinr_rate = usdinr.info.get('regularMarketPrice', 'N/A')
        
        # Get Gold price
        gold = yf.Ticker('GC=F')
        gold_price = gold.info.get('regularMarketPrice', 'N/A')
        
        # Get Bitcoin price
        btc = yf.Ticker('BTC-USD')
        btc_price = btc.info.get('regularMarketPrice', 'N/A')
        
        return JsonResponse({
            'sp500': f"{sp500_price:,.2f} ({sp500_change:+.2f}%)",
            'usd_inr': f"₹{usdinr_rate:,.2f}",
            'gold': f"${gold_price:,.2f}",
            'btc': f"${btc_price:,.2f}"
        })
    except Exception as e:
        return JsonResponse({
            'sp500': 'N/A',
            'usd_inr': 'N/A',
            'gold': 'N/A',
            'btc': 'N/A'
        })

def get_market_news_api(request):
    """API endpoint for market news"""
    news = get_market_news()
    return JsonResponse({'articles': news})

def investment_dashboard(request, id):
    if request.session.has_key('is_logged'):
        user_id = request.session["user_id"]
        if int(id) == 0:
            user = User.objects.get(id=user_id)
            investments = Investment.objects.filter(user=user)
            investment = None
            # Portfolio-wide stats
            total_portfolio_value = sum(inv.current_value for inv in investments)
            total_investments = investments.count()
            total_invested = sum(inv.amount_invested for inv in investments)
            total_returns = sum(inv.current_value - inv.amount_invested for inv in investments)
            total_return_rate = (total_returns / total_invested * 100) if total_invested > 0 else 0
            avg_annualized_return = (
                sum(inv.annualized_return for inv in investments) / total_investments
                if total_investments > 0 else 0
            )
            risk_score = (
                sum(inv.risk_score for inv in investments) / total_investments
                if total_investments > 0 else 0
            )
            types_held = set(inv.investment_type for inv in investments)
            diversification_score = len(types_held) / len(dict(INVESTMENT_TYPE_CHOICES)) * 10 if investments else 0
            diversification_score_percent = diversification_score * 10
            # Recommendations (dummy logic: suggest increase if <10% allocation, decrease if >30%)
            type_distribution = {}
            type_counts = {}
            type_returns = {}
            type_risk = {}
            type_weights = {}
            type_colors = {
                'Stocks': '#4CAF50', 'Mutual Funds': '#2196F3', 'Fixed Deposits': '#FFC107', 'Bonds': '#9C27B0',
                'Real Estate': '#F44336', 'Gold': '#00BCD4', 'Cryptocurrency': '#FF9800', 'Other': '#795548'
            }
            for t, _ in INVESTMENT_TYPE_CHOICES:
                invs = investments.filter(investment_type=t)
                type_counts[t] = invs.count()
                type_distribution[t] = float(sum(inv.current_value for inv in invs))
                type_returns[t] = float(sum(inv.current_value - inv.amount_invested for inv in invs))
                type_risk[t] = sum(inv.risk_score for inv in invs) / invs.count() if invs.count() > 0 else 0
                type_weights[t] = (type_distribution[t] / float(total_portfolio_value) * 100) if total_portfolio_value > 0 else 0
            # Recommendations
            recommendations = []
            for t, _ in INVESTMENT_TYPE_CHOICES:
                current = type_weights[t]
                target = 15  # Example: target 15% for each
                if current < 10:
                    recommendations.append({
                        'asset_type': t, 'action': 'increase', 'difference': target - current,
                        'current_allocation': current, 'target_allocation': target
                    })
                elif current > 30:
                    recommendations.append({
                        'asset_type': t, 'action': 'decrease', 'difference': current - target,
                        'current_allocation': current, 'target_allocation': target
                    })
            # Monthly trends (dummy: use purchase_date month)
            monthly_trends = {}
            for inv in investments:
                month = inv.purchase_date.strftime('%b %Y')
                if month not in monthly_trends:
                    monthly_trends[month] = {'value': 0, 'returns': 0}
                monthly_trends[month]['value'] += float(inv.current_value)
                monthly_trends[month]['returns'] += float(inv.current_value - inv.amount_invested)
            # Sort months
            sorted_months = sorted(monthly_trends.keys(), key=lambda x: datetime.datetime.strptime(x, '%b %Y'))
            labels = sorted_months
            values = [monthly_trends[m]['value'] for m in labels]
            returns = [monthly_trends[m]['returns'] for m in labels]
            # Market news
            news = get_market_news()
            print("DEBUG LABELS:", labels)
            print("DEBUG VALUES:", values)
            context = {
                'investment': investment,
                'total_portfolio_value': total_portfolio_value,
                'total_investments': total_investments,
                'total_returns': total_returns,
                'total_return_rate': total_return_rate,
                'avg_annualized_return': avg_annualized_return,
                'risk_score': risk_score,
                'diversification_score': diversification_score,
                'diversification_score_percent': diversification_score_percent,
                'recommendations': recommendations,
                'type_distribution': type_distribution,
                'type_counts': type_counts,
                'type_returns': type_returns,
                'type_risk': type_risk,
                'type_weights': type_weights,
                'type_colors': type_colors,
                'monthly_trends_labels': labels,
                'monthly_trends_values': [float(v) for v in values],
                'monthly_trends_returns': [float(r) for r in returns],
                'news': news,
            }
            if request.GET.get('modal') == '1':
                return render(request, 'home/investment_dashboard.html', context, using=None, content_type=None, status=None, dirs=None, dictionary=None, request_context=None, template_name=None, context_instance=None, current_app=None, dirs_override=None, using_override=None, block_name='content')
            return render(request, 'home/investment_dashboard.html', context)
        else:
            investment = get_object_or_404(Investment, id=id, user_id=user_id)
            user = investment.user
            investments = Investment.objects.filter(id=investment.id)  # Only this investment
            # Individual investment stats
            total_portfolio_value = investment.current_value
            total_investments = 1
            total_invested = investment.amount_invested
            total_returns = investment.current_value - investment.amount_invested
            total_return_rate = (total_returns / total_invested * 100) if total_invested > 0 else 0
            avg_annualized_return = getattr(investment, 'annualized_return', 0)
            risk_score = getattr(investment, 'risk_score', 0)
            diversification_score = 0
            diversification_score_percent = 0
            # Only this investment's type for type_distribution, etc.
            type_distribution = {investment.get_investment_type_display(): investment.current_value}
            type_counts = {investment.get_investment_type_display(): 1}
            type_returns = {investment.get_investment_type_display(): investment.current_value - investment.amount_invested}
            type_risk = {investment.get_investment_type_display(): getattr(investment, 'risk_score', 0)}
            type_weights = {investment.get_investment_type_display(): 100}
            type_colors = {
                investment.get_investment_type_display(): '#4CAF50'  # You can customize color mapping
            }
            recommendations = []
            # Monthly trends for this investment using value history
            value_history = investment.value_history.order_by('date')
            if value_history.exists():
                labels = [vh.date.strftime('%b %Y') for vh in value_history]
                values = [float(vh.value) for vh in value_history]
            else:
                labels = []
                values = []
            # Always ensure demo data if labels or values are empty
            if not labels or not values:
                import datetime
                today = datetime.date.today()
                labels = []
                values = []
                base_value = float(investment.amount_invested)
                increment = (float(investment.current_value) - base_value) / 5 if base_value != float(investment.current_value) else 0
                for i in range(6):
                    month = (today - datetime.timedelta(days=30 * (5 - i))).strftime('%b %Y')
                    labels.append(month)
                    values.append(base_value + increment * i)
            print("DEBUG - Chart Labels:", labels)
            print("DEBUG - Chart Values:", values)
            returns = [float(v) - float(investment.amount_invested) for v in values]
            # Market news
            news = get_market_news()
            investment_returns = float(investment.current_value) - float(investment.amount_invested)
            print(f"[DEBUG] Investment ID: {investment.id}")
            print(f"[DEBUG] Amount Invested: {investment.amount_invested}")
            print(f"[DEBUG] Current Value: {investment.current_value}")
            print(f"[DEBUG] Returns: {investment_returns}")
            print(f"[DEBUG] Return Rate: {total_return_rate}")
            print(f"[DEBUG] Risk Score: {risk_score}")
        context = {
            'investment': investment,
            'total_portfolio_value': total_portfolio_value,
            'total_investments': total_investments,
            'total_returns': total_returns,
            'total_return_rate': total_return_rate,
            'avg_annualized_return': avg_annualized_return,
            'risk_score': risk_score,
            'diversification_score': diversification_score,
            'diversification_score_percent': diversification_score_percent,
            'recommendations': recommendations,
            'type_distribution': type_distribution,
            'type_counts': type_counts,
            'type_returns': type_returns,
            'type_risk': type_risk,
            'type_weights': type_weights,
            'type_colors': type_colors,
            'monthly_trends_labels': labels,
            'monthly_trends_values': [float(v) for v in values],
            'monthly_trends_returns': [float(r) for r in returns],
            'news': news,
            'investment_returns': investment_returns if 'investment_returns' in locals() else None,
        }
        if request.GET.get('modal') == '1':
            return render(request, 'home/investment_dashboard.html', context, using=None, content_type=None, status=None, dirs=None, dictionary=None, request_context=None, template_name=None, context_instance=None, current_app=None, dirs_override=None, using_override=None, block_name='content')
        return render(request, 'home/investment_dashboard.html', context)
    return redirect('home')

def monthly_trend(request):
    if request.session.has_key('is_logged'):
        user_id = request.session["user_id"]
        user = User.objects.get(id=user_id)
        
        # Get current year
        current_year = datetime.datetime.now().year
        
        # Initialize data structures
        months = []
        income_trend = []
        expense_trend = []
        
        # Get data for each month
        for month in range(1, 13):
            month_name = datetime.date(2000, month, 1).strftime('%b')
            months.append(month_name)
            
            # Get monthly totals
            monthly_data = Addmoney_info.objects.filter(
                user=user,
                Date__year=current_year,
                Date__month=month
            )
            
            income = monthly_data.filter(add_money='Income').aggregate(Sum('quantity'))['quantity__sum'] or 0
            expenses = monthly_data.filter(add_money='Expense').aggregate(Sum('quantity'))['quantity__sum'] or 0
            
            income_trend.append(float(income))
            expense_trend.append(float(expenses))
        
        data = {
            'months': months,
            'income_trend': income_trend,
            'expense_trend': expense_trend
        }
        
        return JsonResponse(data)
    return JsonResponse({'error': 'Not logged in'}, status=401)

@require_http_methods(["GET"])
def transaction_detail(request, transaction_id):
    if not request.session.has_key('is_logged'):
        return JsonResponse({'error': 'Unauthorized'}, status=401)
        
    try:
        user_id = request.session["user_id"]
        transaction = Addmoney_info.objects.get(id=transaction_id, user_id=user_id)
        
        data = {
            'id': transaction.id,
            'add_money': transaction.add_money,
            'quantity': float(transaction.quantity),
            'Category': transaction.Category,
            'Date': transaction.Date.isoformat(),
            'notes': getattr(transaction, 'notes', '')
        }
        
        return JsonResponse(data)
    except Addmoney_info.DoesNotExist:
        return JsonResponse({'error': 'Transaction not found'}, status=404)

@require_http_methods(["DELETE"])
def delete_transaction(request, transaction_id):
    if not request.session.has_key('is_logged'):
        return JsonResponse({'error': 'Unauthorized'}, status=401)
        
    try:
        user_id = request.session["user_id"]
        transaction = Addmoney_info.objects.get(id=transaction_id, user_id=user_id)
        transaction.delete()
        return JsonResponse({'message': 'Transaction deleted successfully'})
    except Addmoney_info.DoesNotExist:
        return JsonResponse({'error': 'Transaction not found'}, status=404)

@require_http_methods(["GET", "POST"])
def edit_transaction(request, transaction_id):
    if not request.session.has_key('is_logged'):
        return redirect('home')
        
    user_id = request.session["user_id"]
    try:
        transaction = Addmoney_info.objects.get(id=transaction_id, user_id=user_id)
        
        if request.method == "POST":
            # Update transaction
            transaction.add_money = request.POST.get('type')
            transaction.quantity = abs(float(request.POST.get('amount') or 0))
            transaction.Category = request.POST.get('category')
            transaction.Date = request.POST.get('date')
            if hasattr(transaction, 'notes'):
                transaction.notes = request.POST.get('notes')
            transaction.save()
            return redirect('tables')
            
        return render(request, 'home/edit_transaction.html', {'transaction': transaction})
    except Addmoney_info.DoesNotExist:
        return redirect('tables')

def export_transactions_csv(request):
    if request.session.has_key('is_logged'):
        user_id = request.session["user_id"]
        user = User.objects.get(id=user_id)
        transactions = Addmoney_info.objects.filter(user=user).order_by('-Date')
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="transactions.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Date', 'Category', 'Type', 'Amount'])
        
        for transaction in transactions:
            writer.writerow([
                transaction.Date.strftime('%Y-%m-%d'),
                transaction.Category,
                transaction.add_money,
                transaction.quantity
            ])
        
        return response
    return JsonResponse({'error': 'Not logged in'}, status=401)

def export_transactions_excel(request):
    if request.session.has_key('is_logged'):
        user_id = request.session["user_id"]
        user = User.objects.get(id=user_id)
        transactions = Addmoney_info.objects.filter(user=user).order_by('-Date')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        # Write headers
        headers = ['Date', 'Category', 'Type', 'Amount']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Write data
        for row, transaction in enumerate(transactions, 2):
            ws.cell(row=row, column=1, value=transaction.Date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=transaction.Category)
            ws.cell(row=row, column=3, value=transaction.add_money)
            ws.cell(row=row, column=4, value=float(transaction.quantity))
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="transactions.xlsx"'
        wb.save(response)
        
        return response
    return JsonResponse({'error': 'Not logged in'}, status=401)

def export_transactions_pdf(request):
    if request.session.has_key('is_logged'):
        user_id = request.session["user_id"]
        user = User.objects.get(id=user_id)
        transactions = Addmoney_info.objects.filter(user=user).order_by('-Date')
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="transactions.pdf"'
        
        # Create PDF document
        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []
        
        # Add title
        styles = getSampleStyleSheet()
        elements.append(Paragraph("Transaction History", styles['Title']))
        elements.append(Spacer(1, 20))
        
        # Create table data
        data = [['Date', 'Category', 'Type', 'Amount']]
        for transaction in transactions:
            data.append([
                transaction.Date.strftime('%Y-%m-%d'),
                transaction.Category,
                transaction.add_money,
                f"₹{abs(transaction.quantity):,.2f}"
            ])
        
        # Create and style table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        
        # Build PDF document
        doc.build(elements)
        return response
    return JsonResponse({'error': 'Not logged in'}, status=401)

@login_required
def spending_forecast(request):
    user = request.user
    today = timezone.now().date()
    
    # Get historical data for the past 12 months
    start_date = today - timedelta(days=365)
    historical_data = Addmoney_info.objects.filter(
        user=user,
        Date__gte=start_date,
        Date__lte=today,
        add_money='Expense'
    ).values('Date', 'Category').annotate(
        daily_total=Sum('quantity')
    ).order_by('Date')
    
    # Prepare data for forecasting
    categories = list(set(item['Category'] for item in historical_data))
    forecast_data = {}
    
    for category in categories:
        # Get category data
        category_data = [item for item in historical_data if item['Category'] == category]
        
        if len(category_data) > 0:
            # Prepare time series data
            dates = [(item['Date'] - start_date).days for item in category_data]
            amounts = [float(item['daily_total']) for item in category_data]
            
            # Reshape data for sklearn
            X = np.array(dates).reshape(-1, 1)
            y = np.array(amounts)
            
            # Create and fit the model
            model = LinearRegression()
            model.fit(X, y)
            
            # Predict next 30 days
            future_dates = np.array(range(366, 396)).reshape(-1, 1)  # Next 30 days after historical data
            predictions = model.predict(future_dates)
            
            # Calculate confidence metrics
            confidence = model.score(X, y)
            trend = 'increasing' if model.coef_[0] > 0 else 'decreasing'
            
            # Store forecast data
            forecast_data[category] = {
                'predictions': [round(max(0, float(p)), 2) for p in predictions],  # Ensure no negative predictions
                'confidence': round(confidence * 100, 2),
                'trend': trend,
                'avg_monthly': round(sum(amounts) / (len(amounts) / 30), 2)
            }
    
    # Calculate overall metrics
    total_expenses = historical_data.aggregate(total=Sum('daily_total'))['total'] or 0
    monthly_avg = total_expenses / 12  # Average monthly expenses
    
    # Get budget information
    budgets = Budget.objects.filter(user=user)
    budget_data = {}
    
    for budget in budgets:
        category_forecast = forecast_data.get(budget.category, {})
        if category_forecast:
            predicted_monthly = sum(category_forecast['predictions']) / 30  # Average daily to monthly
            budget_data[budget.category] = {
                'budget': float(budget.amount),
                'predicted': round(predicted_monthly, 2),
                'variance': round(predicted_monthly - float(budget.amount), 2)
            }
    
    context = {
        'forecast_data': forecast_data,
        'monthly_avg': monthly_avg,
        'budget_data': budget_data,
        'categories': categories,
        'next_month': (today + timedelta(days=30)).strftime('%B %Y')
    }
    
    return render(request, 'home/forecast.html', context)

@require_GET
@login_required
def api_investment_realtime(request):
    import random
    from datetime import timedelta
    user = request.user
    investments = Investment.objects.filter(user=user)
    data = []
    allocation = {}
    performance = []
    today = datetime.date.today()
    # Calculate base portfolio value
    base_value = sum(float(inv.current_value) for inv in investments)
    # Generate fake historical data for the past 7 days
    for i in range(7):
        date = today - timedelta(days=6 - i)
        # Simulate value changes within +/-2%
        value = base_value * (1 + random.uniform(-0.02, 0.02))
        performance.append({'date': date.isoformat(), 'value': value})
    for inv in investments:
        # Update real-time price for stocks
        if inv.stock_symbol:
            inv.update_real_time_price()
        # Portfolio allocation by type
        allocation[inv.investment_type] = allocation.get(inv.investment_type, 0) + float(inv.current_value)
        data.append({
            'id': inv.id,
            'name': inv.name,
            'type': inv.investment_type,
            'stock_symbol': inv.stock_symbol,
            'amount_invested': float(inv.amount_invested),
            'current_value': float(inv.current_value),
            'returns': float(inv.returns),
            'returns_percentage': float(inv.returns_percentage),
            'annualized_return': float(inv.annualized_return),
            'holding_period': inv.holding_period,
            'days_to_maturity': inv.days_to_maturity,
            'status': inv.status,
            'risk_score': inv.risk_score,
            'portfolio_weight': float(inv.portfolio_weight),
            'last_price_update': inv.last_price_update.isoformat() if inv.last_price_update else None,
            'nearing_maturity': inv.days_to_maturity is not None and inv.days_to_maturity <= 7 and inv.days_to_maturity > 0,
        })
    # Portfolio analytics
    total_invested = sum(i['amount_invested'] for i in data)
    total_current = sum(i['current_value'] for i in data)
    total_returns = total_current - total_invested
    total_returns_percentage = (total_returns / total_invested * 100) if total_invested > 0 else 0
    return JsonResponse({
        'investments': data,
        'total_invested': total_invested,
        'total_current': total_current,
        'total_returns': total_returns,
        'total_returns_percentage': total_returns_percentage,
        'allocation': allocation,
        'performance': performance,
    })

@login_required
def export_investments_csv(request):
    user = request.user
    investments = Investment.objects.filter(user=user)
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="investments.csv"'
    writer = csv.writer(response)
    writer.writerow(['Type', 'Name', 'Amount Invested', 'Current Value', 'Returns', 'Return Rate', 'Annualized Return', 'Holding Period', 'Days to Maturity', 'Status'])
    for inv in investments:
        writer.writerow([
            inv.get_investment_type_display(),
            inv.name,
            float(inv.amount_invested),
            float(inv.current_value),
            float(inv.returns),
            float(inv.returns_percentage),
            float(inv.annualized_return),
            inv.holding_period,
            inv.days_to_maturity if inv.days_to_maturity is not None else '-',
            inv.status
        ])
    return response

@login_required
def export_investments_excel(request):
    user = request.user
    investments = Investment.objects.filter(user=user)
    wb = Workbook()
    ws = wb.active
    ws.title = "Investments"
    ws.append(['Type', 'Name', 'Amount Invested', 'Current Value', 'Returns', 'Return Rate', 'Annualized Return', 'Holding Period', 'Days to Maturity', 'Status'])
    for inv in investments:
        ws.append([
            inv.get_investment_type_display(),
            inv.name,
            float(inv.amount_invested),
            float(inv.current_value),
            float(inv.returns),
            float(inv.returns_percentage),
            float(inv.annualized_return),
            inv.holding_period,
            inv.days_to_maturity if inv.days_to_maturity is not None else '-',
            inv.status
        ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="investments.xlsx"'
    return response

@login_required
def export_investments_pdf(request):
    user = request.user
    investments = Investment.objects.filter(user=user)
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = [Paragraph("Investments", styles['Title']), Spacer(1, 12)]
    data = [['Type', 'Name', 'Amount Invested', 'Current Value', 'Returns', 'Return Rate', 'Annualized Return', 'Holding Period', 'Days to Maturity', 'Status']]
    for inv in investments:
        data.append([
            inv.get_investment_type_display(),
            inv.name,
            f"₹{float(inv.amount_invested):,.2f}",
            f"₹{float(inv.current_value):,.2f}",
            f"₹{float(inv.returns):,.2f}",
            f"{float(inv.returns_percentage):.2f}%",
            f"{float(inv.annualized_return):.2f}%",
            f"{inv.holding_period} days",
            inv.days_to_maturity if inv.days_to_maturity is not None else '-',
            inv.status
        ])
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#23272f')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#353b48')),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')

def add_transaction(request):
    return render(request, 'home/add_transaction.html')

@login_required
def edit_budget(request, budget_id):
    budget = get_object_or_404(Budget, id=budget_id, user=request.user)
    if request.method == 'POST':
        budget.name = request.POST.get('budgetName')
        budget.amount = request.POST.get('budgetAmount')
        budget.category = request.POST.get('budgetCategory')
        budget.period = request.POST.get('budgetPeriod')
        budget.save()
        return redirect('budgets')
    return render(request, 'home/edit_budget.html', {'budget': budget})

@login_required
def delete_budget(request, budget_id):
    budget = get_object_or_404(Budget, id=budget_id, user=request.user)
    budget.delete()
    return redirect('budgets')

def investment_dashboard_full(request):
    # This will show the dashboard for all investments of the logged-in user
    return investment_dashboard(request, id=0)

def export_investment_csv(request, id):
    investment = get_object_or_404(Investment, id=id, user_id=request.session["user_id"])
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="investment_{investment.id}.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'Name', 'Type', 'Amount Invested', 'Current Value', 'Returns', 'Return Rate', 'Risk Score',
        'Purchase Date', 'Maturity Date', 'Status', 'Notes', 'Sector', 'Dividend Yield', 'PE Ratio', 'Quantity'
    ])
    writer.writerow([
        investment.name,
        investment.get_investment_type_display(),
        float(investment.amount_invested),
        float(investment.current_value),
        float(investment.current_value - investment.amount_invested),
        investment.returns_percentage,
        investment.risk_score,
        investment.purchase_date,
        investment.maturity_date or '',
        investment.get_status_display(),
        investment.notes or '',
        investment.sector or '',
        investment.dividend_yield or '',
        investment.pe_ratio or '',
        investment.quantity,
    ])
    return response
     